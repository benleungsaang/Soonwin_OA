import json
import os
import glob
from openpyxl import Workbook, load_workbook
from decimal import Decimal

# 定义需要单独提取的核心字段（映射关系：Excel列名 -> JSON字段名）
CORE_FIELDS = {
    "model": "Model",
    "original_model": "OriginalModel",
    "machine_weight": "MachineWeight",
    "dimensions": "Dimensions",
    "general_power": "GeneralPower",
    "power_supply": "PowerSupply",
    "show_price": "ShowPrice",
    "original_price": "OriginalPrice",
    "machine_type": "machine_type",
    "remark": "remark",
    "brand": "brand"
}

def clear_screen():
    """清屏（兼容Windows和Linux/Mac）"""
    os.system('cls' if os.name == 'nt' else 'clear')

def get_file_list(file_ext):
    """
    获取当前目录下指定后缀的文件列表
    :param file_ext: 文件后缀（如 .json .xlsx）
    :return: 文件列表（绝对路径）
    """
    file_pattern = f"*{file_ext}"
    file_list = glob.glob(file_pattern)
    # 按文件名排序
    file_list.sort()
    return file_list

def select_file(file_ext):
    """
    交互式选择文件
    :param file_ext: 文件后缀
    :return: 选中的文件路径，None表示取消
    """
    file_list = get_file_list(file_ext)
    if not file_list:
        print(f"\n当前目录下未找到{file_ext}文件！")
        input("按回车键返回主菜单...")
        return None

    print(f"\n=== 可选的{file_ext}文件 ===")
    for idx, file in enumerate(file_list, 1):
        print(f"{idx}. {file}")
    print("0. 取消返回")

    while True:
        try:
            choice = input("\n请输入文件序号选择（0-取消）：").strip()
            if not choice:
                continue
            choice = int(choice)
            if choice == 0:
                return None
            elif 1 <= choice <= len(file_list):
                return file_list[choice-1]
            else:
                print(f"输入无效！请输入1-{len(file_list)}之间的数字")
        except ValueError:
            print("输入无效！请输入数字序号")

def json_to_xlsx(json_file_path, xlsx_file_path=None):
    """
    将JSON文件转换为XLSX文件
    :param json_file_path: 输入的JSON文件路径
    :param xlsx_file_path: 输出的XLSX文件路径（默认自动生成）
    """
    # 自动生成输出文件名
    if not xlsx_file_path:
        xlsx_file_path = os.path.splitext(json_file_path)[0] + ".xlsx"

    # 1. 读取JSON数据
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        if not isinstance(json_data, list):
            raise ValueError("JSON数据必须是数组格式")
    except Exception as e:
        print(f"\n❌ 读取JSON文件失败：{e}")
        input("按回车键返回...")
        return

    # 2. 创建Excel工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "设备数据"

    # 3. 写入表头（核心字段 + custom_attrs）
    headers = list(CORE_FIELDS.keys()) + ["custom_attrs"]
    ws.append(headers)

    # 4. 处理每条数据并写入
    success_count = 0
    for item in json_data:
        row_data = []
        # 提取核心字段
        for excel_field, json_field in CORE_FIELDS.items():
            # 处理Decimal类型（价格字段）
            value = item.get(json_field, "")
            if excel_field in ["show_price", "original_price"]:
                value = Decimal(str(value)) if value else Decimal("0.00")
            row_data.append(value)

        # 提取非核心字段，存入custom_attrs（JSON字符串）
        custom_attrs = {}
        for key, value in item.items():
            if key not in CORE_FIELDS.values():
                custom_attrs[key] = value
        row_data.append(json.dumps(custom_attrs, ensure_ascii=False, indent=2))

        # 写入行数据
        ws.append(row_data)
        success_count += 1

    # 5. 保存Excel文件
    try:
        wb.save(xlsx_file_path)
        print(f"\n✅ JSON转XLSX成功！")
        print(f"📄 输入文件：{json_file_path}")
        print(f"📄 输出文件：{xlsx_file_path}")
        print(f"📊 处理数据条数：{success_count}")
    except Exception as e:
        print(f"\n❌ 保存XLSX文件失败：{e}")

    input("\n按回车键返回主菜单...")

def xlsx_to_json(xlsx_file_path, json_file_path=None):
    """
    将XLSX文件转换为JSON文件（还原原始格式）
    :param xlsx_file_path: 输入的XLSX文件路径
    :param json_file_path: 输出的JSON文件路径（默认自动生成）
    """
    # 自动生成输出文件名
    if not json_file_path:
        json_file_path = os.path.splitext(xlsx_file_path)[0] + "_restored.json"

    # 1. 加载Excel文件
    try:
        wb = load_workbook(xlsx_file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"\n❌ 读取XLSX文件失败：{e}")
        input("按回车键返回...")
        return

    # 2. 读取表头并验证
    headers = [cell.value for cell in ws[1]]
    required_headers = list(CORE_FIELDS.keys()) + ["custom_attrs"]
    missing_headers = [h for h in required_headers if h not in headers]
    if missing_headers:
        print(f"\n❌ Excel表头不符合要求！缺少字段：{', '.join(missing_headers)}")
        input("按回车键返回...")
        return

    # 3. 处理每行数据，还原JSON格式
    json_data = []
    success_count = 0
    fail_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if not row[0]:  # 跳过空行
            continue

        # 构建单条设备数据
        item = {}
        row_dict = dict(zip(headers, row))

        # 还原核心字段
        for excel_field, json_field in CORE_FIELDS.items():
            value = row_dict.get(excel_field, "")
            # 处理数值类型转换
            if excel_field in ["show_price", "original_price"]:
                value = float(value) if value else 0.0
            elif excel_field == "machine_type":
                value = int(value) if value else 0
            item[json_field] = value

        # 还原custom_attrs中的字段
        custom_attrs_str = row_dict.get("custom_attrs", "{}")
        try:
            custom_attrs = json.loads(custom_attrs_str) if custom_attrs_str else {}
            item.update(custom_attrs)
            json_data.append(item)
            success_count += 1
        except json.JSONDecodeError as e:
            print(f"\n⚠️  解析第{success_count + fail_count + 1}行custom_attrs失败：{e}，该行已跳过")
            fail_count += 1
            continue

    # 4. 保存JSON文件
    try:
        with open(json_file_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=4)
        print(f"\n✅ XLSX转JSON成功！")
        print(f"📄 输入文件：{xlsx_file_path}")
        print(f"📄 输出文件：{json_file_path}")
        print(f"📊 成功处理：{success_count} 条 | 失败跳过：{fail_count} 条")
    except Exception as e:
        print(f"\n❌ 保存JSON文件失败：{e}")

    input("\n按回车键返回主菜单...")

def main_menu():
    """主菜单"""
    while True:
        clear_screen()
        print("="*40)
        print("     JSON ↔ XLSX 转换工具 v2.0")
        print("="*40)
        print("请选择操作：")
        print("1. JSON文件 转 XLSX文件")
        print("2. XLSX文件 转 JSON文件")
        print("0. 退出程序")
        print("="*40)

        # 选择操作
        while True:
            choice = input("\n请输入操作序号（0-2）：").strip()
            if choice in ["0", "1", "2"]:
                break
            print("输入无效！请输入0、1或2")

        if choice == "0":
            print("\n👋 程序已退出！")
            break
        elif choice == "1":
            # JSON转XLSX
            print("\n--- 选择要转换的JSON文件 ---")
            json_file = select_file(".json")
            if json_file:
                json_to_xlsx(json_file)
        elif choice == "2":
            # XLSX转JSON
            print("\n--- 选择要转换的XLSX文件 ---")
            xlsx_file = select_file(".xlsx")
            if xlsx_file:
                xlsx_to_json(xlsx_file)

if __name__ == "__main__":
    try:
        main_menu()
    except KeyboardInterrupt:
        print("\n\n👋 程序被用户中断，已退出！")
    except Exception as e:
        print(f"\n❌ 程序运行出错：{e}")
        input("按回车键退出...")