import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime, date, timedelta
import calendar
import random
import uuid
import json
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ---------------------- 全局配置与枚举 ----------------------
fake = Faker("zh_CN")

# 操作类型枚举
class OperationType:
    LEAVE = "leave"  # 请假
    OVERTIME = "overtime"  # 加班
    MAKE_UP = "make_up"  # 补卡
    APPEAL = "appeal"  # 迟到/早退申诉
    BUSINESS_TRIP = "business_trip"  # 出差
    ADJUST = "adjust"  # 管理员手动调整

# 操作状态枚举
class OperationStatus:
    DRAFT = "draft"  # 草稿
    SUBMITTED = "submitted"  # 已提交
    APPROVING = "approving"  # 审批中
    APPROVED = "approved"  # 审批通过
    REJECTED = "rejected"  # 审批驳回
    CANCELLED = "cancelled"  # 已撤销

# ---------------------- 数据模型（简化版，无需数据库） ----------------------
class PunchRecord:
    def __init__(self, emp_id, name, punch_type, punch_time, inner_ip=None, device_id=None):
        self.emp_id = emp_id
        self.name = name
        self.punch_type = punch_type  # 上班/下班
        self.punch_time = punch_time
        self.inner_ip = inner_ip
        self.device_id = device_id

class Employee:
    def __init__(self, emp_id, name, base_salary=5000):
        self.emp_id = emp_id
        self.name = name
        self.base_salary = base_salary  # 月工资，默认5000

class AttendanceOperation:
    def __init__(self, emp_id, name, operation_type, operation_status, start_time, end_time, duration, reason,
                 approver_id=None, approver_name=None, approve_time=None, approve_opinion=None):
        self.id = str(uuid.uuid4())
        self.emp_id = emp_id
        self.name = name
        self.operation_type = operation_type
        self.operation_status = operation_status
        self.start_time = start_time
        self.end_time = end_time
        self.duration = duration
        self.reason = reason
        self.approver_id = approver_id
        self.approver_name = approver_name
        self.approve_time = approve_time
        self.approve_opinion = approve_opinion
        self.create_time = datetime.now()
        self.update_time = datetime.now()

# ---------------------- 核心工具函数 ----------------------
def get_default_work_dates(year, month):
    """获取当月默认上班日期（排除周日）"""
    work_dates = []
    total_days = calendar.monthrange(year, month)[1]
    for day in range(1, total_days + 1):
        current_date = date(year, month, day)
        if current_date.weekday() != 6:  # 排除周日
            work_dates.append(current_date)
    return work_dates

def generate_test_employees(count=5):
    """生成测试员工数据"""
    employees = []
    for i in range(count):
        emp_id = f"{1000 + i:04d}"  # 工号：1000,1001...
        name = fake.name()
        employees.append(Employee(emp_id=emp_id, name=name, base_salary=5000))
    return employees

def generate_test_data(employees, year, month):
    """为员工生成当月测试考勤数据（修复日期越界问题）"""
    punch_records = []
    attendance_ops = []
    total_days = calendar.monthrange(year, month)[1]
    # 生成当月有效日期列表（排除周日）
    valid_days = [d for d in range(1, total_days+1) if date(year, month, d).weekday() != 6]

    for emp in employees:
        emp_id = emp.emp_id
        name = emp.name

        # 1. 生成打卡记录
        for day in range(1, total_days + 1):
            current_date = date(year, month, day)
            # 周日不打卡
            if current_date.weekday() == 6:
                continue

            # 85%概率正常打卡，15%漏打卡
            if random.random() < 0.85:
                # 上班打卡时间（8:00-9:00）
                work_start = datetime(year, month, day, 8 + random.randint(0, 1), random.randint(0, 59))
                punch_records.append(PunchRecord(
                    emp_id=emp_id,
                    name=name,
                    punch_type="上班",
                    punch_time=work_start,
                    inner_ip=f"192.168.1.{random.randint(1, 254)}",
                    device_id=f"DEVICE_{random.randint(1000, 9999)}"
                ))

                # 下班打卡时间（17:00-18:00）
                work_end = datetime(year, month, day, 17 + random.randint(0, 1), random.randint(0, 59))
                punch_records.append(PunchRecord(
                    emp_id=emp_id,
                    name=name,
                    punch_type="下班",
                    punch_time=work_end,
                    inner_ip=f"192.168.1.{random.randint(1, 254)}",
                    device_id=f"DEVICE_{random.randint(1000, 9999)}"
                ))

        # 2. 生成补卡记录（2条，审批通过）- 修复日期越界
        if len(valid_days) >= 2:
            missed_days = random.sample(valid_days, 2)
        else:
            missed_days = valid_days[:1] if valid_days else []

        for day in missed_days:
            # 确保日期在当月范围内
            if 1 <= day <= total_days:
                ops = AttendanceOperation(
                    emp_id=emp_id,
                    name=name,
                    operation_type=OperationType.MAKE_UP,
                    operation_status=OperationStatus.APPROVED,
                    start_time=datetime(year, month, day, 0, 0),
                    end_time=datetime(year, month, day, 23, 59),
                    duration=1.0,
                    reason="忘记打卡，申请补卡",
                    approver_id="9999",
                    approver_name="管理员",
                    approve_time=datetime(year, month, min(day + 1, total_days), 10, 0),
                    approve_opinion="同意补卡"
                )
                attendance_ops.append(ops)

        # 3. 生成出差记录（3天，审批通过）- 修复日期越界
        if len(valid_days) >= 3:
            trip_start = random.choice([d for d in valid_days if d + 2 <= total_days])
            trip_end = min(trip_start + 2, total_days)
            trip_op = AttendanceOperation(
                emp_id=emp_id,
                name=name,
                operation_type=OperationType.BUSINESS_TRIP,
                operation_status=OperationStatus.APPROVED,
                start_time=datetime(year, month, trip_start, 0, 0),
                end_time=datetime(year, month, trip_end, 23, 59),
                duration=float(trip_end - trip_start + 1),
                reason=f"{fake.city()}出差对接项目",
                approver_id="9999",
                approver_name="管理员",
                approve_time=datetime(year, month, max(trip_start - 1, 1), 15, 0),
                approve_opinion="同意出差"
            )
            attendance_ops.append(trip_op)

        # 4. 生成请假记录（1-2天，审批通过）- 修复日期越界
        if valid_days:
            leave_day = random.choice([d for d in valid_days if d + 1 <= total_days])
            leave_days = random.choice([1, 2]) if leave_day + 1 <= total_days else 1
            leave_end = min(leave_day + leave_days - 1, total_days)

            leave_op = AttendanceOperation(
                emp_id=emp_id,
                name=name,
                operation_type=OperationType.LEAVE,
                operation_status=OperationStatus.APPROVED,
                start_time=datetime(year, month, leave_day, 0, 0),
                end_time=datetime(year, month, leave_end, 23, 59),
                duration=float(leave_end - leave_day + 1),
                reason=fake.sentence(nb_words=5),
                approver_id="9999",
                approver_name="管理员",
                approve_time=datetime(year, month, max(leave_day - 1, 1), 11, 0),
                approve_opinion="同意请假"
            )
            attendance_ops.append(leave_op)

    return punch_records, attendance_ops

def is_valid_punch_day(punch_records, emp_id, target_date):
    """判断员工某一天打卡是否有效"""
    emp_punch = [r for r in punch_records if r.emp_id == emp_id and r.punch_time.date() == target_date]
    if not emp_punch:
        return False
    has_work = any(r.punch_type == "上班" for r in emp_punch)
    has_off = any(r.punch_type == "下班" for r in emp_punch)
    return has_work and has_off

def get_approved_ops(attendance_ops, emp_id, year, month):
    """获取员工当月审批通过的考勤操作"""
    start = datetime(year, month, 1)
    end = datetime(year, month, calendar.monthrange(year, month)[1], 23, 59, 59)
    ops = [op for op in attendance_ops if op.emp_id == emp_id and op.operation_status == OperationStatus.APPROVED
           and op.start_time >= start and op.end_time <= end]
    return ops

def calculate_attendance(employees, punch_records, attendance_ops, year, month, work_dates):
    """计算考勤和工资"""
    results = []
    trip_allowance_per_day = 120

    for emp in employees:
        emp_id = emp.emp_id
        name = emp.name
        base_salary = emp.base_salary  # 使用员工的月工资
        should_work_days = len(work_dates)
        # 计算日工资（按应出勤天数）
        daily_salary = base_salary / should_work_days if should_work_days > 0 else 0

        # 1. 获取员工审批通过的操作
        approved_ops = get_approved_ops(attendance_ops, emp_id, year, month)

        # 2. 统计各类操作
        make_up_ops = [op for op in approved_ops if op.operation_type == OperationType.MAKE_UP]
        trip_ops = [op for op in approved_ops if op.operation_type == OperationType.BUSINESS_TRIP]
        leave_ops = [op for op in approved_ops if op.operation_type == OperationType.LEAVE]

        # 3. 计算实际有效上班天数
        actual_work_days = 0
        trip_days_total = 0.0
        leave_days_total = 0.0

        for work_date in work_dates:
            # 正常打卡有效
            if is_valid_punch_day(punch_records, emp_id, work_date):
                actual_work_days += 1
                continue

            # 补卡有效
            if any(op.start_time.date() == work_date for op in make_up_ops):
                actual_work_days += 1
                continue

            # 出差有效
            trip_ops_for_day = [op for op in trip_ops if op.start_time.date() <= work_date <= op.end_time.date()]
            if trip_ops_for_day:
                actual_work_days += 1
                trip_days_total += 1
                continue

            # 请假不计入
            leave_ops_for_day = [op for op in leave_ops if op.start_time.date() <= work_date <= op.end_time.date()]
            if leave_ops_for_day:
                leave_days_total += 1
                continue

        # 4. 计算扣款和补贴
        absent_days = max(0, should_work_days - actual_work_days)
        leave_deduct = absent_days * daily_salary
        trip_allowance = trip_days_total * trip_allowance_per_day
        # 出差时的工资也计入
        trip_salary = trip_days_total * daily_salary
        total_salary = base_salary - leave_deduct + trip_allowance + trip_salary

        results.append({
            "emp_id": emp_id,
            "name": name,
            "year": year,
            "month": month,
            "should_work_days": should_work_days,
            "actual_work_days": actual_work_days,
            "absent_days": absent_days,
            "leave_days": leave_days_total,
            "trip_days": trip_days_total,
            "daily_salary": round(daily_salary, 2),
            "leave_deduct": round(leave_deduct, 2),
            "trip_allowance": round(trip_allowance, 2),
            "trip_salary": round(trip_salary, 2),
            "base_salary": base_salary,
            "total_salary": round(total_salary, 2)
        })

    return results

def export_to_excel(results, file_path):
    """导出统计结果到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "月度考勤工资统计"

    # 表头
    headers = [
        "员工工号", "员工姓名", "年份", "月份", "基础工资（元）", "应上班天数", "实际有效天数",
        "缺勤天数", "请假天数", "日工资（元）", "出差天数", "出差补贴（元）", "出差工资（元）", "缺勤扣款（元）",
        "应发工资（元）"
    ]

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

    # 写入数据
    for row, res in enumerate(results, 2):
        data = [
            res["emp_id"], res["name"], res["year"], res["month"], res["base_salary"],
            res["should_work_days"], res["actual_work_days"], res["absent_days"],
            res["leave_days"], res["daily_salary"], res["trip_days"], res["trip_allowance"],
            res["trip_salary"], res["leave_deduct"], res["total_salary"]
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

    # 调整列宽
    col_widths = [10, 10, 8, 8, 12, 10, 10, 8, 8, 10, 8, 12, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(file_path)

# ---------------------- TK界面实现 ----------------------
class AttendanceStatApp:
    def __init__(self, root):
        self.root = root
        self.root.title("👨‍💼 员工考勤工资统计工具")
        self.root.geometry("1100x750")

        # 全局变量
        self.year = tk.IntVar(value=datetime.now().year)
        self.month = tk.IntVar(value=datetime.now().month)
        self.work_dates = get_default_work_dates(self.year.get(), self.month.get())  # 选中的上班日期
        self.employees = []
        self.punch_records = []
        self.attendance_ops = []
        self.stat_results = []

        # 初始化界面
        self._create_widgets()

    def _create_widgets(self):
        # 1. 月份选择区
        frame_month = ttk.LabelFrame(self.root, text="📅 选择月份")
        frame_month.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_month, text="年份：").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(frame_month, textvariable=self.year, width=10).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_month, text="月份：").grid(row=0, column=2, padx=5, pady=5)
        ttk.Entry(frame_month, textvariable=self.month, width=10).grid(row=0, column=3, padx=5, pady=5)

        ttk.Button(frame_month, text="🔄 加载默认上班日期", command=self._load_default_dates).grid(row=0, column=4, padx=5, pady=5)

        # 2. 日期选择区
        frame_dates = ttk.LabelFrame(self.root, text="🗓️ 选择当月上班日期（默认排除周日，可手动调整）")
        frame_dates.pack(fill="x", padx=10, pady=5)

        # 日期选择框架
        self.date_frame = ttk.Frame(frame_dates)
        self.date_frame.pack(fill="x", padx=5, pady=5)

        # 日期选择变量
        self.date_vars = []

        # 3. 操作按钮区
        frame_buttons = ttk.Frame(self.root)
        frame_buttons.pack(fill="x", padx=10, pady=5)

        ttk.Button(frame_buttons, text="🧑‍🤝‍🧑 生成测试数据", command=self._generate_test_data).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="💰 设置员工工资", command=self._set_employee_salary).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="🧮 计算考勤工资", command=self._calculate_stat).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="📤 导出Excel", command=self._export_excel).pack(side="left", padx=5)

        # 4. 结果展示区
        frame_result = ttk.LabelFrame(self.root, text="📊 统计结果")
        frame_result.pack(fill="both", expand=True, padx=10, pady=5)

        # 结果表格
        self.result_tree = ttk.Treeview(frame_result, columns=(
            "emp_id", "name", "base_salary", "should", "actual", "absent", "daily_salary", "trip_days", "trip_allowance", "trip_salary", "deduct", "total"
        ), show="headings")

        # 设置表头（带emoji）
        self.result_tree.heading("emp_id", text="🆔 员工工号")
        self.result_tree.heading("name", text="👤 员工姓名")
        self.result_tree.heading("base_salary", text="💴 基础工资（元）")
        self.result_tree.heading("should", text="📅 应上班天数")
        self.result_tree.heading("actual", text="✅ 实际有效天数")
        self.result_tree.heading("absent", text="❌ 缺勤天数")
        self.result_tree.heading("daily_salary", text="💵 日工资（元）")
        self.result_tree.heading("trip_days", text="✈️ 出差天数")
        self.result_tree.heading("trip_allowance", text="🎁 出差补贴（元）")
        self.result_tree.heading("trip_salary", text="💸 出差工资（元）")
        self.result_tree.heading("deduct", text="🚫 缺勤扣款（元）")
        self.result_tree.heading("total", text="💰 应发工资（元）")

        # 设置列宽
        self.result_tree.column("emp_id", width=80)
        self.result_tree.column("name", width=80)
        self.result_tree.column("base_salary", width=90)
        self.result_tree.column("should", width=90)
        self.result_tree.column("actual", width=90)
        self.result_tree.column("absent", width=80)
        self.result_tree.column("daily_salary", width=90)
        self.result_tree.column("trip_days", width=80)
        self.result_tree.column("trip_allowance", width=100)
        self.result_tree.column("trip_salary", width=100)
        self.result_tree.column("deduct", width=100)
        self.result_tree.column("total", width=100)

        # 滚动条
        scrollbar = ttk.Scrollbar(frame_result, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.result_tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)

    def _load_default_dates(self):
        """加载默认上班日期并显示在按周排列的多选框中"""
        try:
            year = self.year.get()
            month = self.month.get()
            if month < 1 or month > 12:
                messagebox.showerror("❌ 错误", "月份必须是1-12！")
                return

            # 获取当月所有日期
            total_days = calendar.monthrange(year, month)[1]
            all_dates = [date(year, month, day) for day in range(1, total_days + 1)]

            # 清空日期框架
            for widget in self.date_frame.winfo_children():
                widget.destroy()

            # 重新创建日期选择控件
            self.date_vars = []

            # 创建星期标题行（周一开头）
            week_header_frame = ttk.Frame(self.date_frame)
            week_header_frame.pack(fill="x", padx=5, pady=2)
            ttk.Label(week_header_frame, text="一", width=8).pack(side="left", padx=1)
            ttk.Label(week_header_frame, text="二", width=8).pack(side="left", padx=1)
            ttk.Label(week_header_frame, text="三", width=8).pack(side="left", padx=1)
            ttk.Label(week_header_frame, text="四", width=8).pack(side="left", padx=1)
            ttk.Label(week_header_frame, text="五", width=8).pack(side="left", padx=1)
            ttk.Label(week_header_frame, text="六", width=8).pack(side="left", padx=1)
            ttk.Label(week_header_frame, text="日", width=8).pack(side="left", padx=1)

            # 按周创建日期选择行，以周一开头
            current_date = date(year, month, 1)
            first_day_weekday = current_date.weekday()  # 0是周一，6是周日
            start_of_week = current_date - timedelta(days=first_day_weekday)

            end_of_month = date(year, month, total_days)
            current_week_start = start_of_week

            while current_week_start <= end_of_month:
                week_frame = ttk.Frame(self.date_frame)
                week_frame.pack(fill="x", padx=5, pady=2)

                # 创建一周的日期选择
                for i in range(7):  # 一周七天，从周一开始
                    day_date = current_week_start + timedelta(days=i)
                    day_str = day_date.strftime("%d")

                    # 检查日期是否在当前月份内
                    if day_date.month == month:
                        # 创建复选框
                        var = tk.BooleanVar()
                        # 默认选中非周日的日期
                        if day_date.weekday() != 6:  # 不是周日
                            var.set(True)

                        chk = tk.Checkbutton(week_frame, text=day_str, variable=var, width=6)
                        chk.pack(side="left", padx=1)

                        # 记录日期和对应的变量
                        self.date_vars.append((day_date, var))
                    else:
                        # 不是当前月的日期，显示为空白
                        ttk.Label(week_frame, text="", width=8).pack(side="left", padx=1)

                # 移动到下周
                current_week_start += timedelta(days=7)

            messagebox.showinfo("✅ 成功", f"已加载{year}年{month}月默认上班日期 📅")
        except Exception as e:
            messagebox.showerror("❌ 错误", f"加载日期失败：{str(e)}")

    def _generate_test_data(self):
        """生成测试数据（修复日期越界错误）"""
        try:
            year = self.year.get()
            month = self.month.get()
            # 清空原有数据，避免重复
            self.employees = []
            self.punch_records = []
            self.attendance_ops = []

            # 生成5名员工
            self.employees = generate_test_employees(5)
            # 生成考勤数据
            self.punch_records, self.attendance_ops = generate_test_data(self.employees, year, month)

            # 先计算预览数据
            preview_work_dates = [d for d, v in self.date_vars if v.get()] if self.date_vars else get_default_work_dates(year, month)
            self.preview_stat = calculate_attendance(
                self.employees, self.punch_records, self.attendance_ops,
                year, month, preview_work_dates
            )

            messagebox.showinfo("✅ 成功", "已生成5名员工的测试考勤数据 🧑‍🤝‍🧑！")
            # 显示生成的测试数据
            self._show_generated_data()
        except Exception as e:
            messagebox.showerror("❌ 错误", f"生成数据失败：{str(e)}")

    def _show_generated_data(self):
        """展示完整的测试数据（包含考勤统计，支持编辑）"""
        if not self.employees:
            return

        # 创建新窗口展示数据
        data_window = tk.Toplevel(self.root)
        data_window.title("📋 生成的测试数据")
        data_window.geometry("1200x700")

        # 创建框架
        main_frame = ttk.Frame(data_window)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 创建Notebook用于分页显示不同类型的数据
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill="both", expand=True)

        # 第一页：员工基本信息
        emp_frame = ttk.Frame(notebook)
        notebook.add(emp_frame, text="👤 员工信息")

        emp_tree = ttk.Treeview(emp_frame, columns=("emp_id", "name", "base_salary"), show="headings")
        emp_tree.heading("emp_id", text="🆔 员工工号")
        emp_tree.heading("name", text="👤 员工姓名")
        emp_tree.heading("base_salary", text="💴 基础工资（元）")

        # 设置列宽
        emp_tree.column("emp_id", width=120)
        emp_tree.column("name", width=150)
        emp_tree.column("base_salary", width=120)

        emp_tree.pack(fill="both", expand=True, pady=5)

        # 添加员工数据到Treeview
        for emp in self.employees:
            emp_tree.insert("", tk.END, values=(emp.emp_id, emp.name, emp.base_salary), tags=(emp.emp_id,))

        # 双击编辑员工信息
        def edit_employee(event):
            item = emp_tree.selection()[0]
            tags = emp_tree.item(item, "tags")
            if not tags:
                return
            emp_id = tags[0]
            current_vals = emp_tree.item(item, "values")

            # 创建编辑窗口
            edit_win = tk.Toplevel(data_window)
            edit_win.title(f"✏️ 编辑员工 - {emp_id}")
            edit_win.geometry("350x200")
            edit_win.transient(data_window)
            edit_win.grab_set()

            # 员工工号（不可编辑）
            ttk.Label(edit_win, text="🆔 员工工号：").grid(row=0, column=0, padx=10, pady=10, sticky="w")
            ttk.Label(edit_win, text=emp_id).grid(row=0, column=1, padx=10, pady=10, sticky="w")

            # 员工姓名
            ttk.Label(edit_win, text="👤 员工姓名：").grid(row=1, column=0, padx=10, pady=10, sticky="w")
            name_var = tk.StringVar(value=current_vals[1])
            name_entry = ttk.Entry(edit_win, textvariable=name_var, width=20)
            name_entry.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

            # 基础工资
            ttk.Label(edit_win, text="💴 基础工资：").grid(row=2, column=0, padx=10, pady=10, sticky="w")
            salary_var = tk.DoubleVar(value=float(current_vals[2]))
            salary_entry = ttk.Entry(edit_win, textvariable=salary_var, width=20)
            salary_entry.grid(row=2, column=1, padx=10, pady=10, sticky="ew")

            # 保存按钮
            def save_emp():
                new_name = name_var.get().strip()
                new_salary = salary_var.get()
                if not new_name:
                    messagebox.showwarning("⚠️ 警告", "姓名不能为空！")
                    return

                # 更新数据
                for emp in self.employees:
                    if emp.emp_id == emp_id:
                        emp.name = new_name
                        emp.base_salary = new_salary
                        break

                # 更新表格
                emp_tree.item(item, values=(emp_id, new_name, new_salary))
                edit_win.destroy()

            btn_frame = ttk.Frame(edit_win)
            btn_frame.grid(row=3, column=0, columnspan=2, pady=10)
            ttk.Button(btn_frame, text="💾 保存", command=save_emp).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="❌ 取消", command=edit_win.destroy).pack(side="left", padx=5)

        emp_tree.bind("<Double-1>", edit_employee)

        # 第二页：打卡记录
        punch_frame = ttk.Frame(notebook)
        notebook.add(punch_frame, text="📝 打卡记录")

        punch_tree = ttk.Treeview(punch_frame, columns=("emp_id", "name", "punch_type", "punch_time", "inner_ip", "device_id"), show="headings")
        punch_tree.heading("emp_id", text="🆔 员工工号")
        punch_tree.heading("name", text="👤 员工姓名")
        punch_tree.heading("punch_type", text="📌 打卡类型")
        punch_tree.heading("punch_time", text="🕒 打卡时间")
        punch_tree.heading("inner_ip", text="🌐 内网IP")
        punch_tree.heading("device_id", text="🖥️ 设备ID")

        # 设置列宽
        punch_tree.column("emp_id", width=100)
        punch_tree.column("name", width=120)
        punch_tree.column("punch_type", width=100)
        punch_tree.column("punch_time", width=180)
        punch_tree.column("inner_ip", width=120)
        punch_tree.column("device_id", width=120)

        # 滚动条
        punch_scroll = ttk.Scrollbar(punch_frame, orient="vertical", command=punch_tree.yview)
        punch_tree.configure(yscrollcommand=punch_scroll.set)
        punch_scroll.pack(side="right", fill="y")
        punch_tree.pack(side="left", fill="both", expand=True, pady=5)

        # 添加打卡记录到Treeview
        for idx, record in enumerate(self.punch_records):
            punch_tree.insert("", tk.END, values=(
                record.emp_id, record.name, record.punch_type,
                record.punch_time.strftime("%Y-%m-%d %H:%M:%S"), record.inner_ip, record.device_id
            ), tags=(str(idx),))

        # 双击编辑打卡记录
        def edit_punch(event):
            item = punch_tree.selection()[0]
            idx = int(punch_tree.item(item, "tags")[0])
            record = self.punch_records[idx]
            current_vals = punch_tree.item(item, "values")

            edit_win = tk.Toplevel(data_window)
            edit_win.title(f"✏️ 编辑打卡记录 - {record.emp_id}")
            edit_win.geometry("400x300")
            edit_win.transient(data_window)
            edit_win.grab_set()

            # 打卡类型
            ttk.Label(edit_win, text="📌 打卡类型：").grid(row=0, column=0, padx=10, pady=8, sticky="w")
            type_var = tk.StringVar(value=current_vals[2])
            type_combo = ttk.Combobox(edit_win, textvariable=type_var, values=["上班", "下班"], width=18)
            type_combo.grid(row=0, column=1, padx=10, pady=8, sticky="w")

            # 打卡时间
            ttk.Label(edit_win, text="🕒 打卡时间：").grid(row=1, column=0, padx=10, pady=8, sticky="w")
            time_var = tk.StringVar(value=current_vals[3])
            time_entry = ttk.Entry(edit_win, textvariable=time_var, width=20)
            time_entry.grid(row=1, column=1, padx=10, pady=8, sticky="w")

            # IP地址
            ttk.Label(edit_win, text="🌐 内网IP：").grid(row=2, column=0, padx=10, pady=8, sticky="w")
            ip_var = tk.StringVar(value=current_vals[4])
            ip_entry = ttk.Entry(edit_win, textvariable=ip_var, width=20)
            ip_entry.grid(row=2, column=1, padx=10, pady=8, sticky="w")

            # 设备ID
            ttk.Label(edit_win, text="🖥️ 设备ID：").grid(row=3, column=0, padx=10, pady=8, sticky="w")
            device_var = tk.StringVar(value=current_vals[5])
            device_entry = ttk.Entry(edit_win, textvariable=device_var, width=20)
            device_entry.grid(row=3, column=1, padx=10, pady=8, sticky="w")

            def save_punch():
                try:
                    # 验证时间格式
                    new_time = datetime.strptime(time_var.get(), "%Y-%m-%d %H:%M:%S")
                    # 更新记录
                    record.punch_type = type_var.get()
                    record.punch_time = new_time
                    record.inner_ip = ip_var.get()
                    record.device_id = device_var.get()
                    # 更新表格
                    punch_tree.item(item, values=(
                        record.emp_id, record.name, record.punch_type,
                        record.punch_time.strftime("%Y-%m-%d %H:%M:%S"),
                        record.inner_ip, record.device_id
                    ))
                    edit_win.destroy()
                except ValueError:
                    messagebox.showwarning("⚠️ 警告", "时间格式错误！请使用：YYYY-MM-DD HH:MM:SS")

            btn_frame = ttk.Frame(edit_win)
            btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
            ttk.Button(btn_frame, text="💾 保存", command=save_punch).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="❌ 取消", command=edit_win.destroy).pack(side="left", padx=5)

        punch_tree.bind("<Double-1>", edit_punch)

        # 第二页：考勤操作
        ops_frame = ttk.Frame(notebook)
        notebook.add(ops_frame, text="📑 考勤操作")

        ops_tree = ttk.Treeview(ops_frame, columns=(
            "emp_id", "name", "operation_type", "operation_status", "start_time", "end_time", "duration", "reason"
        ), show="headings")
        ops_tree.heading("emp_id", text="🆔 员工工号")
        ops_tree.heading("name", text="👤 员工姓名")
        ops_tree.heading("operation_type", text="📌 操作类型")
        ops_tree.heading("operation_status", text="📊 操作状态")
        ops_tree.heading("start_time", text="🕒 开始时间")
        ops_tree.heading("end_time", text="🕒 结束时间")
        ops_tree.heading("duration", text="⏱️ 时长(天)")
        ops_tree.heading("reason", text="📝 原因")

        # 设置列宽
        ops_tree.column("emp_id", width=100)
        ops_tree.column("name", width=120)
        ops_tree.column("operation_type", width=120)
        ops_tree.column("operation_status", width=120)
        ops_tree.column("start_time", width=150)
        ops_tree.column("end_time", width=150)
        ops_tree.column("duration", width=100)
        ops_tree.column("reason", width=200)

        # 滚动条
        ops_scroll = ttk.Scrollbar(ops_frame, orient="vertical", command=ops_tree.yview)
        ops_tree.configure(yscrollcommand=ops_scroll.set)
        ops_scroll.pack(side="right", fill="y")
        ops_tree.pack(side="left", fill="both", expand=True, pady=5)

        # 添加考勤操作到Treeview
        for idx, op in enumerate(self.attendance_ops):
            ops_tree.insert("", tk.END, values=(
                op.emp_id, op.name, op.operation_type, op.operation_status,
                op.start_time.strftime("%Y-%m-%d %H:%M:%S"), op.end_time.strftime("%Y-%m-%d %H:%M:%S"),
                op.duration, op.reason
            ), tags=(str(idx),))

        # 双击编辑考勤操作
        def edit_ops(event):
            item = ops_tree.selection()[0]
            idx = int(ops_tree.item(item, "tags")[0])
            op = self.attendance_ops[idx]
            current_vals = ops_tree.item(item, "values")

            edit_win = tk.Toplevel(data_window)
            edit_win.title(f"✏️ 编辑考勤操作 - {op.emp_id}")
            edit_win.geometry("450x400")
            edit_win.transient(data_window)
            edit_win.grab_set()

            # 操作类型
            ttk.Label(edit_win, text="📌 操作类型：").grid(row=0, column=0, padx=10, pady=8, sticky="w")
            type_var = tk.StringVar(value=current_vals[2])
            type_combo = ttk.Combobox(edit_win, textvariable=type_var,
                                     values=["leave", "overtime", "make_up", "appeal", "business_trip", "adjust"],
                                     width=18)
            type_combo.grid(row=0, column=1, padx=10, pady=8, sticky="w")

            # 操作状态
            ttk.Label(edit_win, text="📊 操作状态：").grid(row=1, column=0, padx=10, pady=8, sticky="w")
            status_var = tk.StringVar(value=current_vals[3])
            status_combo = ttk.Combobox(edit_win, textvariable=status_var,
                                       values=["draft", "submitted", "approving", "approved", "rejected", "cancelled"],
                                       width=18)
            status_combo.grid(row=1, column=1, padx=10, pady=8, sticky="w")

            # 开始时间
            ttk.Label(edit_win, text="🕒 开始时间：").grid(row=2, column=0, padx=10, pady=8, sticky="w")
            start_var = tk.StringVar(value=current_vals[4])
            start_entry = ttk.Entry(edit_win, textvariable=start_var, width=25)
            start_entry.grid(row=2, column=1, padx=10, pady=8, sticky="w")

            # 结束时间
            ttk.Label(edit_win, text="🕒 结束时间：").grid(row=3, column=0, padx=10, pady=8, sticky="w")
            end_var = tk.StringVar(value=current_vals[5])
            end_entry = ttk.Entry(edit_win, textvariable=end_var, width=25)
            end_entry.grid(row=3, column=1, padx=10, pady=8, sticky="w")

            # 时长
            ttk.Label(edit_win, text="⏱️ 时长(天)：").grid(row=4, column=0, padx=10, pady=8, sticky="w")
            dur_var = tk.DoubleVar(value=float(current_vals[6]))
            dur_entry = ttk.Entry(edit_win, textvariable=dur_var, width=20)
            dur_entry.grid(row=4, column=1, padx=10, pady=8, sticky="w")

            # 原因
            ttk.Label(edit_win, text="📝 原因：").grid(row=5, column=0, padx=10, pady=8, sticky="nw")
            reason_var = tk.StringVar(value=current_vals[7])
            reason_text = tk.Text(edit_win, width=30, height=3)
            reason_text.insert("1.0", reason_var.get())
            reason_text.grid(row=5, column=1, padx=10, pady=8, sticky="w")

            def save_ops():
                try:
                    # 验证时间格式
                    new_start = datetime.strptime(start_var.get(), "%Y-%m-%d %H:%M:%S")
                    new_end = datetime.strptime(end_var.get(), "%Y-%m-%d %H:%M:%S")
                    # 更新记录
                    op.operation_type = type_var.get()
                    op.operation_status = status_var.get()
                    op.start_time = new_start
                    op.end_time = new_end
                    op.duration = dur_var.get()
                    op.reason = reason_text.get("1.0", tk.END).strip()
                    # 更新表格
                    ops_tree.item(item, values=(
                        op.emp_id, op.name, op.operation_type, op.operation_status,
                        op.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                        op.end_time.strftime("%Y-%m-%d %H:%M:%S"),
                        op.duration, op.reason
                    ))
                    edit_win.destroy()
                except ValueError as e:
                    messagebox.showwarning("⚠️ 警告", f"时间格式错误！{str(e)}")

            btn_frame = ttk.Frame(edit_win)
            btn_frame.grid(row=6, column=0, columnspan=2, pady=10)
            ttk.Button(btn_frame, text="💾 保存", command=save_ops).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="❌ 取消", command=edit_win.destroy).pack(side="left", padx=5)

        ops_tree.bind("<Double-1>", edit_ops)

        # 第四页：考勤统计预览
        stat_frame = ttk.Frame(notebook)
        notebook.add(stat_frame, text="📊 考勤统计预览")

        stat_tree = ttk.Treeview(stat_frame, columns=(
            "emp_id", "name", "should", "actual", "absent", "leave", "trip", "daily", "deduct", "allowance", "total"
        ), show="headings")
        stat_tree.heading("emp_id", text="🆔 员工工号")
        stat_tree.heading("name", text="👤 员工姓名")
        stat_tree.heading("should", text="📅 应上班天数")
        stat_tree.heading("actual", text="✅ 实际天数")
        stat_tree.heading("absent", text="❌ 缺勤天数")
        stat_tree.heading("leave", text="🏥 请假天数")
        stat_tree.heading("trip", text="✈️ 出差天数")
        stat_tree.heading("daily", text="💵 日工资")
        stat_tree.heading("deduct", text="🚫 缺勤扣款")
        stat_tree.heading("allowance", text="🎁 出差补贴")
        stat_tree.heading("total", text="💰 应发工资")

        # 设置列宽
        stat_tree.column("emp_id", width=100)
        stat_tree.column("name", width=120)
        stat_tree.column("should", width=90)
        stat_tree.column("actual", width=90)
        stat_tree.column("absent", width=90)
        stat_tree.column("leave", width=90)
        stat_tree.column("trip", width=90)
        stat_tree.column("daily", width=90)
        stat_tree.column("deduct", width=100)
        stat_tree.column("allowance", width=100)
        stat_tree.column("total", width=100)

        # 滚动条
        stat_scroll = ttk.Scrollbar(stat_frame, orient="vertical", command=stat_tree.yview)
        stat_tree.configure(yscrollcommand=stat_scroll.set)
        stat_scroll.pack(side="right", fill="y")
        stat_tree.pack(side="left", fill="both", expand=True, pady=5)

        # 添加统计数据到Treeview
        for idx, stat in enumerate(self.preview_stat):
            stat_tree.insert("", tk.END, values=(
                stat["emp_id"], stat["name"], stat["should_work_days"],
                stat["actual_work_days"], stat["absent_days"], stat["leave_days"],
                stat["trip_days"], stat["daily_salary"], stat["leave_deduct"],
                stat["trip_allowance"], stat["total_salary"]
            ), tags=(str(idx),))

        # 双击编辑统计数据（仅允许编辑数值型字段）
        def edit_stat(event):
            item = stat_tree.selection()[0]
            idx = int(stat_tree.item(item, "tags")[0])
            stat = self.preview_stat[idx]
            current_vals = stat_tree.item(item, "values")

            edit_win = tk.Toplevel(data_window)
            edit_win.title(f"✏️ 编辑考勤统计 - {stat['emp_id']}")
            edit_win.geometry("400x400")
            edit_win.transient(data_window)
            edit_win.grab_set()

            # 应上班天数
            ttk.Label(edit_win, text="📅 应上班天数：").grid(row=0, column=0, padx=10, pady=8, sticky="w")
            should_var = tk.IntVar(value=int(current_vals[2]))
            should_entry = ttk.Entry(edit_win, textvariable=should_var, width=15)
            should_entry.grid(row=0, column=1, padx=10, pady=8, sticky="w")

            # 实际有效天数
            ttk.Label(edit_win, text="✅ 实际天数：").grid(row=1, column=0, padx=10, pady=8, sticky="w")
            actual_var = tk.IntVar(value=int(current_vals[3]))
            actual_entry = ttk.Entry(edit_win, textvariable=actual_var, width=15)
            actual_entry.grid(row=1, column=1, padx=10, pady=8, sticky="w")

            # 缺勤天数
            ttk.Label(edit_win, text="❌ 缺勤天数：").grid(row=2, column=0, padx=10, pady=8, sticky="w")
            absent_var = tk.IntVar(value=int(current_vals[4]))
            absent_entry = ttk.Entry(edit_win, textvariable=absent_var, width=15)
            absent_entry.grid(row=2, column=1, padx=10, pady=8, sticky="w")

            # 请假天数
            ttk.Label(edit_win, text="🏥 请假天数：").grid(row=3, column=0, padx=10, pady=8, sticky="w")
            leave_var = tk.FloatVar(value=float(current_vals[5]))
            leave_entry = ttk.Entry(edit_win, textvariable=leave_var, width=15)
            leave_entry.grid(row=3, column=1, padx=10, pady=8, sticky="w")

            # 出差天数
            ttk.Label(edit_win, text="✈️ 出差天数：").grid(row=4, column=0, padx=10, pady=8, sticky="w")
            trip_var = tk.FloatVar(value=float(current_vals[6]))
            trip_entry = ttk.Entry(edit_win, textvariable=trip_var, width=15)
            trip_entry.grid(row=4, column=1, padx=10, pady=8, sticky="w")

            # 日工资
            ttk.Label(edit_win, text="💵 日工资：").grid(row=5, column=0, padx=10, pady=8, sticky="w")
            daily_var = tk.FloatVar(value=float(current_vals[7]))
            daily_entry = ttk.Entry(edit_win, textvariable=daily_var, width=15)
            daily_entry.grid(row=5, column=1, padx=10, pady=8, sticky="w")

            # 缺勤扣款
            ttk.Label(edit_win, text="🚫 缺勤扣款：").grid(row=6, column=0, padx=10, pady=8, sticky="w")
            deduct_var = tk.FloatVar(value=float(current_vals[8]))
            deduct_entry = ttk.Entry(edit_win, textvariable=deduct_var, width=15)
            deduct_entry.grid(row=6, column=1, padx=10, pady=8, sticky="w")

            # 出差补贴
            ttk.Label(edit_win, text="🎁 出差补贴：").grid(row=7, column=0, padx=10, pady=8, sticky="w")
            allowance_var = tk.FloatVar(value=float(current_vals[9]))
            allowance_entry = ttk.Entry(edit_win, textvariable=allowance_var, width=15)
            allowance_entry.grid(row=7, column=1, padx=10, pady=8, sticky="w")

            def save_stat():
                try:
                    # 更新统计数据
                    stat["should_work_days"] = should_var.get()
                    stat["actual_work_days"] = actual_var.get()
                    stat["absent_days"] = absent_var.get()
                    stat["leave_days"] = leave_var.get()
                    stat["trip_days"] = trip_var.get()
                    stat["daily_salary"] = daily_var.get()
                    stat["leave_deduct"] = deduct_var.get()
                    stat["trip_allowance"] = allowance_var.get()
                    # 重新计算应发工资
                    stat["total_salary"] = round(stat["base_salary"] - stat["leave_deduct"] + stat["trip_allowance"] + (stat["trip_days"] * stat["daily_salary"]), 2)

                    # 更新表格
                    stat_tree.item(item, values=(
                        stat["emp_id"], stat["name"], stat["should_work_days"],
                        stat["actual_work_days"], stat["absent_days"], stat["leave_days"],
                        stat["trip_days"], stat["daily_salary"], stat["leave_deduct"],
                        stat["trip_allowance"], stat["total_salary"]
                    ))
                    edit_win.destroy()
                except ValueError as e:
                    messagebox.showwarning("⚠️ 警告", f"输入格式错误！{str(e)}")

            btn_frame = ttk.Frame(edit_win)
            btn_frame.grid(row=8, column=0, columnspan=2, pady=10)
            ttk.Button(btn_frame, text="💾 保存", command=save_stat).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="❌ 取消", command=edit_win.destroy).pack(side="left", padx=5)

        stat_tree.bind("<Double-1>", edit_stat)

        # 关闭按钮
        close_btn = ttk.Button(main_frame, text="❌ 关闭", command=data_window.destroy)
        close_btn.pack(pady=5)

    def _set_employee_salary(self):
        """设置员工工资"""
        if not self.employees:
            messagebox.warning("⚠️ 提示", "请先生成测试数据 🧑‍🤝‍🧑！")
            return

        # 创建新窗口设置员工工资
        salary_window = tk.Toplevel(self.root)
        salary_window.title("💰 设置员工工资")
        salary_window.geometry("400x300")

        # 创建Treeview显示员工信息
        tree = ttk.Treeview(salary_window, columns=("emp_id", "name", "base_salary"), show="headings")
        tree.heading("emp_id", text="🆔 员工工号")
        tree.heading("name", text="👤 员工姓名")
        tree.heading("base_salary", text="💴 月工资")

        # 设置列宽
        tree.column("emp_id", width=100)
        tree.column("name", width=120)
        tree.column("base_salary", width=120)

        tree.pack(fill="both", expand=True, padx=10, pady=10)

        # 添加员工数据到Treeview
        for emp in self.employees:
            tree.insert("", tk.END, values=(emp.emp_id, emp.name, emp.base_salary), tags=(emp.emp_id,))

        # 双击编辑工资
        def edit_salary(event):
            item = tree.selection()[0]
            emp_id = tree.item(item, "tags")[0]
            current_salary = tree.item(item, "values")[2]

            # 弹出输入框修改工资
            new_salary = tk.simpledialog.askfloat("💴 修改工资", f"员工工号: {emp_id}\n请输入新工资：", initialvalue=float(current_salary))
            if new_salary is not None and new_salary > 0:
                # 更新员工工资
                for emp in self.employees:
                    if emp.emp_id == emp_id:
                        emp.base_salary = new_salary
                        break

                # 更新Treeview显示
                tree.item(item, values=(emp_id, [emp.name for emp in self.employees if emp.emp_id == emp_id][0], new_salary))

        tree.bind("<Double-1>", edit_salary)

        # 确认按钮
        ttk.Button(salary_window, text="✅ 确认", command=salary_window.destroy).pack(pady=5)

    def _calculate_stat(self):
        """计算考勤和工资"""
        if not self.employees or not self.punch_records:
            messagebox.warning("⚠️ 提示", "请先生成测试数据 🧑‍🤝‍🧑！")
            return

        try:
            # 获取选中的上班日期
            selected_dates = []
            for date_obj, var in self.date_vars:
                if var.get():  # 如果复选框被选中
                    selected_dates.append(date_obj)

            if not selected_dates:
                messagebox.warning("⚠️ 提示", "请至少选择1个上班日期 📅！")
                return

            year = self.year.get()
            month = self.month.get()

            # 计算统计结果
            self.stat_results = calculate_attendance(
                self.employees, self.punch_records, self.attendance_ops,
                year, month, selected_dates
            )

            # 清空结果表格
            for item in self.result_tree.get_children():
                self.result_tree.delete(item)

            # 插入结果
            for res in self.stat_results:
                self.result_tree.insert("", tk.END, values=(
                    res["emp_id"], res["name"], res["base_salary"], res["should_work_days"],
                    res["actual_work_days"], res["absent_days"], res["daily_salary"],
                    res["trip_days"], res["trip_allowance"], res["trip_salary"],
                    res["leave_deduct"], res["total_salary"]
                ))

            messagebox.showinfo("✅ 成功", "考勤工资计算完成 🧮！")
        except Exception as e:
            messagebox.showerror("❌ 错误", f"计算失败：{str(e)}")

    def _export_excel(self):
        """导出Excel"""
        if not self.stat_results:
            messagebox.warning("⚠️ 提示", "请先计算考勤工资 🧮！")
            return

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件 📊", "*.xlsx"), ("所有文件", "*.*")],
                title="📤 保存统计结果"
            )
            if file_path:
                export_to_excel(self.stat_results, file_path)
                messagebox.showinfo("✅ 成功", f"已导出至：{file_path} 📤")
        except Exception as e:
            messagebox.showerror("❌ 错误", f"导出失败：{str(e)}")

# ---------------------- 运行程序 ----------------------
if __name__ == "__main__":
    # 安装依赖提示（首次运行需执行）
    # pip install faker openpyxl
    root = tk.Tk()
    app = AttendanceStatApp(root)
    root.mainloop()