import os
import json
import pandas as pd
from datetime import datetime

def get_video_files(directory):
    """获取目录中的所有视频文件"""
    video_extensions = ['.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv', '.webm', '.m4v']
    video_files = []
    
    for file in os.listdir(directory):
        file_ext = os.path.splitext(file)[1].lower()
        if file_ext in video_extensions:
            video_files.append(file)
    
    return video_files

def generate_video_data(video_files, directory):
    """生成视频数据，只填写title和filePath，其他字段保留但为空"""
    videos = []
    
    for file in video_files:
        # 获取不带后缀的文件名作为标题
        title = os.path.splitext(file)[0]
        # 真实的文件路径
        file_path = os.path.join(directory, file).replace("\\", "/")  # 使用正斜杠路径
        
        # 为每个视频创建完整的字段结构，但只填写title和filePath，其他保留空值
        video_data = {
            "id": "",
            "title": title,
            "filePath": file_path,
            "fileSize": 0,
            "tagIds": [],
            "remark": "",
            "uploadTime": "",
            "duration": 0,
            "thumbnailPath": "",
            "transcode": "",
            "tagNames": []
        }
        videos.append(video_data)
    
    return videos

def create_json_file(videos, output_path):
    """创建JSON文件"""
    data = {"videos": videos}
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON文件已生成: {output_path}")

def create_excel_file(videos, output_path):
    """创建Excel文件，只保留title, tagNames, remark字段（按指定顺序），并将所有单元格设置为文本格式"""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # 只选择需要的字段
    selected_fields = []
    for video in videos:
        selected_video = {
            "title": video.get("title", ""),
            "tagNames": ', '.join(video.get("tagNames", [])),  # 将tagNames数组转换为逗号分隔的字符串
            "remark": video.get("remark", "")
        }
        selected_fields.append(selected_video)
    
    # 按照指定顺序创建DataFrame
    df = pd.DataFrame(selected_fields)[["title", "tagNames", "remark"]]
    
    # 使用openpyxl创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "视频数据"
    
    # 将DataFrame数据添加到工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 使用批量设置方式将所有数据单元格设置为文本格式（提高效率）
    max_row = ws.max_row
    if max_row > 1:  # 如果有数据行（不包括标题行）
        # 获取数据区域（跳过标题行，从第2行开始）
        target_range = ws[f'A2:C{max_row}']  # 假设最多10000行，实际使用max_row
        for row in target_range:
            for cell in row:
                cell.number_format = '@'  # '@' 表示文本格式
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已生成: {output_path}")

def json_to_excel(json_path, excel_path):
    """将JSON文件转换为Excel文件，只保留title, tagNames, remark字段，并将所有单元格设置为文本格式"""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    videos = data.get("videos", [])
    
    # 只选择需要的字段
    selected_fields = []
    for video in videos:
        selected_video = {
            "title": video.get("title", ""),
            "tagNames": ', '.join(video.get("tagNames", [])),  # 将tagNames数组转换为逗号分隔的字符串
            "remark": video.get("remark", "")
        }
        selected_fields.append(selected_video)
    
    # 按照指定顺序创建DataFrame
    df = pd.DataFrame(selected_fields)[["title", "tagNames", "remark"]]
    
    # 使用openpyxl创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "视频数据"
    
    # 将DataFrame数据添加到工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 使用批量设置方式将所有数据单元格设置为文本格式（提高效率）
    max_row = ws.max_row
    if max_row > 1:  # 如果有数据行（不包括标题行）
        # 获取数据区域（跳过标题行，从第2行开始）
        target_range = ws[f'A2:C{max_row}']  # 假设最多10000行，实际使用max_row
        for row in target_range:
            for cell in row:
                cell.number_format = '@'  # '@' 表示文本格式
    
    # 保存文件
    wb.save(excel_path)
    print(f"JSON已转换为Excel: {excel_path}")

def excel_to_json(excel_path, json_path):
    """将Excel文件转换为JSON文件"""
    df = pd.read_excel(excel_path)
    videos = df.to_dict('records')
    
    # 重新获取当前目录中的视频文件，以便填入filePath
    current_dir = os.path.dirname(excel_path)
    video_files = get_video_files(current_dir)
    
    # 创建文件名到完整路径的映射
    file_path_map = {}
    for file in video_files:
        title = os.path.splitext(file)[0]  # 获取不带扩展名的文件名
        full_path = os.path.join(current_dir, file).replace("\\", "/")
        file_path_map[title] = full_path
    
    # 将Excel数据转换回JSON格式
    processed_videos = []
    for video in videos:
        # 创建完整的视频数据结构，用Excel中的值填充特定字段
        processed_video = {
            "id": "",
            "title": "",
            "filePath": "",  # 在转换或生成JSON时再填入
            "fileSize": 0,
            "tagIds": [],  # 保持为空数组
            "remark": "",
            "uploadTime": "",
            "duration": 0,
            "thumbnailPath": "",
            "transcode": "",
            "tagNames": []
        }
        
        # 处理可能的NaN值和数据类型
        title = video.get("title", "")
        if not pd.isna(title):
            processed_video["title"] = str(title)  # 确保转换为字符串
            # 从映射中获取对应的文件路径
            if str(title) in file_path_map:
                processed_video["filePath"] = file_path_map[str(title)]
        
        remark = video.get("remark", "")
        if not pd.isna(remark):
            processed_video["remark"] = str(remark)  # 确保转换为字符串
        else:
            processed_video["remark"] = ""  # 如果是NaN，设置为空字符串
        
        # 处理tagNames：将逗号分隔的字符串转换回数组
        # 先将中文逗号替换为英文逗号，提高代码健壮性
        tagNames_str = video.get("tagNames", "")
        if pd.isna(tagNames_str):
            processed_video["tagNames"] = []
        elif isinstance(tagNames_str, (int, float)):
            # 如果tagNames是数字（如123456.0），转换为字符串再处理
            tagNames_str = str(tagNames_str)
            if tagNames_str.endswith('.0'):
                tagNames_str = tagNames_str[:-2]  # 去掉.0后缀
            processed_video["tagNames"] = [tagNames_str]
        elif isinstance(tagNames_str, str) and tagNames_str.strip() != "":
            # 将中文逗号替换为英文逗号
            tagNames_str = tagNames_str.replace('，', ',')
            # 按逗号分割字符串，去除空格，并过滤空字符串
            processed_video["tagNames"] = [tag.strip() for tag in tagNames_str.split(",") if tag.strip()]
        else:
            processed_video["tagNames"] = []
        
        processed_videos.append(processed_video)
    
    data = {"videos": processed_videos}
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Excel已转换为JSON: {json_path}")

def confirm_action(prompt):
    """确认操作的函数"""
    response = input(f"{prompt} (y/n, 直接回车默认为y): ").strip().lower()
    return response in ['', 'y', 'yes']

def main():
    while True:
        print("\n" + "="*50)
        print("视频批量上传处理工具")
        print("="*50)
        print("1. 生成JSON文件 (从当前目录视频文件)")
        print("2. 生成Excel文件 (从当前目录视频文件)")
        print("3. JSON文件转换为Excel")
        print("4. Excel文件转换为JSON")
        print("5. 退出")
        print("-"*50)
        
        choice = input("请选择要执行的任务 (输入序号): ").strip()
        
        if choice == '5':
            print("退出程序。")
            break
        elif choice in ['1', '2', '3', '4']:
            current_dir = os.getcwd()
            
            if choice in ['1', '2']:
                # 获取视频文件列表
                video_files = get_video_files(current_dir)
                print(f"\n在当前目录找到 {len(video_files)} 个视频文件:")
                for i, file in enumerate(video_files, 1):
                    print(f"  {i}. {file}")
                
                if not video_files:
                    print("未找到任何视频文件，无法执行任务。")
                    continue
                
                if not confirm_action(f"确认要处理这 {len(video_files)} 个视频文件吗?"):
                    print("操作已取消。")
                    continue
                
                # 生成视频数据
                videos = generate_video_data(video_files, current_dir)
                
                if choice == '1':
                    # 生成JSON文件
                    json_output = os.path.join(current_dir, "视频批量上传.json")
                    if confirm_action(f"确认要生成JSON文件到: {json_output}?"):
                        create_json_file(videos, json_output)
                        print("JSON文件生成完成！")
                    else:
                        print("操作已取消。")
                
                elif choice == '2':
                    # 生成Excel文件
                    excel_output = os.path.join(current_dir, "视频批量上传.xlsx")
                    if confirm_action(f"确认要生成Excel文件到: {excel_output}?"):
                        create_excel_file(videos, excel_output)
                        print("Excel文件生成完成！")
                    else:
                        print("操作已取消。")
            
            elif choice == '3':
                # JSON转Excel
                json_files = [f for f in os.listdir(current_dir) if f.endswith('.json')]
                print(f"\n在当前目录找到 {len(json_files)} 个JSON文件:")
                for i, file in enumerate(json_files, 1):
                    print(f"  {i}. {file}")
                
                if not json_files:
                    print("未找到任何JSON文件，无法执行任务。")
                    continue
                
                json_idx = input(f"请选择要转换的JSON文件 (输入序号 1-{len(json_files)}, 直接回车选择第一个): ").strip()
                if json_idx == '':
                    json_idx = '1'
                
                try:
                    json_idx = int(json_idx) - 1
                    if 0 <= json_idx < len(json_files):
                        selected_json = json_files[json_idx]
                        json_path = os.path.join(current_dir, selected_json)
                        
                        excel_output = os.path.join(current_dir, f"{os.path.splitext(selected_json)[0]}.xlsx")
                        
                        if confirm_action(f"确认要将 '{selected_json}' 转换为 '{os.path.basename(excel_output)}'?"):
                            json_to_excel(json_path, excel_output)
                            print("JSON到Excel转换完成！")
                        else:
                            print("操作已取消。")
                    else:
                        print("无效的序号。")
                        continue
                except ValueError:
                    print("请输入有效的数字。")
                    continue
            
            elif choice == '4':
                # Excel转JSON
                excel_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]
                print(f"\n在当前目录找到 {len(excel_files)} 个Excel文件:")
                for i, file in enumerate(excel_files, 1):
                    print(f"  {i}. {file}")
                
                if not excel_files:
                    print("未找到任何Excel文件，无法执行任务。")
                    continue
                
                excel_idx = input(f"请选择要转换的Excel文件 (输入序号 1-{len(excel_files)}, 直接回车选择第一个): ").strip()
                if excel_idx == '':
                    excel_idx = '1'
                
                try:
                    excel_idx = int(excel_idx) - 1
                    if 0 <= excel_idx < len(excel_files):
                        selected_excel = excel_files[excel_idx]
                        excel_path = os.path.join(current_dir, selected_excel)
                        
                        json_output = os.path.join(current_dir, f"{os.path.splitext(selected_excel)[0]}.json")
                        
                        if confirm_action(f"确认要将 '{selected_excel}' 转换为 '{os.path.basename(json_output)}'?"):
                            excel_to_json(excel_path, json_output)
                            print("Excel到JSON转换完成！")
                        else:
                            print("操作已取消。")
                    else:
                        print("无效的序号。")
                        continue
                except ValueError:
                    print("请输入有效的数字。")
                    continue
        else:
            print("无效的选项，请重新选择。")

if __name__ == "__main__":
    main()